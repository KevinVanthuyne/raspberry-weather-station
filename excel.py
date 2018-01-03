from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

import os.path

class Excel:
    """Class for writing weather readings to an excel file"""

    def __init__(self, excel_file_name):
        self.excel_file_name = excel_file_name
        # create an excel file if there is none
        if not os.path.isfile(excel_file_name):
            self.workbook = self.create_excel(excel_file_name)
            print("Excel log file created")
        else:
            self.workbook = load_workbook(excel_file_name)
            print("Excel log file found")

    def create_excel(self, excel_file_name):
        """ Create an excel file to log weather readings in """
        wb = Workbook()  # create workbook
        ws = wb.active  # select worksheet
        ws.title = "Temperature"  # set worksheet title
        # set first row (titles) to bold
        row = ws.row_dimensions[1]
        row.font = Font(bold = True)
        # fill in titles
        ws.cell(row = 1, column = 1, value = "Day")
        ws.cell(row = 1, column = 2, value = "Hour")
        ws.cell(row = 1, column = 3, value = "Weather hour")
        ws.cell(row = 1, column = 4, value = "Outside temperature (°C)")
        ws.cell(row = 1, column = 5, value = "Inside temperature (°C)")
        ws.cell(row = 1, column = 6, value = "Inside humidity (%)")

        # set column widths
        ws.column_dimensions['A'].width = 12  # Day
        ws.column_dimensions['B'].width = 10  # Hour
        ws.column_dimensions['C'].width = 10  # Weather hour
        ws.column_dimensions['D'].width = 10  # Outside temperature
        ws.column_dimensions['E'].width = 10  # Inside temperature
        ws.column_dimensions['F'].width = 10  # Inside humidity

        # save the file
        wb.save(excel_file_name)

        return wb

    def write_to_excel(self, day, hour, weather_hour, outside_temp, inside_temp, inside_humid):
        sheet = self.workbook.active
        # get next unused row
        row = sheet.max_row + 1
        # write data to colums in row
        sheet.cell(row = row, column = 1, value = day)
        sheet.cell(row = row, column = 2, value = hour)
        sheet.cell(row = row, column = 3, value = weather_hour)
        sheet.cell(row = row, column = 4, value = outside_temp)
        sheet.cell(row = row, column = 5, value = inside_temp)
        sheet.cell(row = row, column = 6, value = inside_humid)
        # save the file
        self.workbook.save(self.excel_file_name)

        print("Written to excel. Row {}".format(row))
        print("------------------------------------")
