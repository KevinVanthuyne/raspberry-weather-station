# coding=utf-8
# !/usr/bin/env python3

# Library used:
# https://github.com/csparpa/pyowm

from luma.led_matrix.device import max7219
from luma.core.interface.serial import spi, noop
from luma.core.render import canvas
from luma.core.legacy import text
from luma.core.legacy.font import proportional, TINY_FONT

from pyowm import OWM

import Adafruit_DHT as dht

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

import RPi.GPIO as GPIO

import time
import os.path
import datetime

__excel_file_name = '/home/pi/Programs/Weerstation/temperatures.xlsx'
__DHT22_pin = 15  # GPIO pin on Raspberry Pi for DHT22 sensor (BCM number)
__lightsensor_pin = 27  #GPIO pin connected to digital out of LM393 light sensor

def run():
    # Matrix setup
    serial = spi(port = 0, device = 0, gpio = noop())
    device = max7219(serial)
    device.contrast(0)

    # Weather setup
    API_key = '80393a1060ab07030ec77f53770a9760'
    owm = OWM(API_key)

    # Excel setup
    workbook = load_workbook(__excel_file_name)

    # Light sensor setup
    GPIO.setmode(GPIO.BCM)
    GPIO.setup(__lightsensor_pin, GPIO.IN)

    # Show X on screen to inform Pi is booting
    with canvas(device) as draw:
        text(draw, (0, 0), "X", fill="white", font=proportional(TINY_FONT))

    # Main loop
    while True:
        now = datetime.datetime.now()
        print("It's now: {}".format(datetime.datetime.now().time()))
        # Get current weather info for Leuven
        # obs = owm.weather_at_coords(50.875494, 4.711183) # Observation
        # Peizegem city:
        obs = owm.weather_at_coords(50.978978, 4.211429)

        # get outside weather
        weather = obs.get_weather()
        timestamp = weather.get_reference_time()
        day = datetime.datetime.fromtimestamp(timestamp).strftime('%d/%m/%Y')
        hour = datetime.datetime.fromtimestamp(timestamp).strftime('%H:%M:%S')
        outside_temp = weather.get_temperature(unit = 'celsius')['temp']

        # get inside sensor readings
        inside_humid, inside_temp = dht.read_retry(dht.DHT22, __DHT22_pin, retries = 5, delay_seconds = 1)
        # if a reading could be gotten
        if inside_humid is not None and inside_temp is not None:
            inside_humid = round(inside_humid, 2)
            inside_temp = round(inside_temp, 2)
        else:
            inside_temp = ""
            inside_humid = ""

        # print some data
        print("Fetched data is from: {} {}".format(day, hour))
        print("it's now {}°C".format(outside_temp))
        print("Short weather status: {}".format(weather.get_status()))
        print("Detailed weather status: {}".format(weather.get_detailed_status()))
        print("Sunrise: {}".format(weather.get_sunrise_time('iso')))
        print("Sunset: {}".format(weather.get_sunset_time('iso')))
        print("")
        print("Inside temp: {}°C".format(inside_temp))
        print("Inside humidity: {}%".format(inside_humid))
        print("")

        # log data to Excel
        write_to_excel(workbook, day, hour, outside_temp, inside_temp, inside_humid)

        # If it's not dark in the room
        # (Sensor output = 1 if dark, 0 if light)
        if not GPIO.input(__lightsensor_pin):
            # Display current temperature
            with canvas(device) as draw:
                text(draw, (0, 0), str(outside_temp), fill="white", font=proportional(TINY_FONT))
        else:
            print("Device cleared")
            device.clear()  # TODO device.show() and hide()

        # Sleep 5 minutes
        time.sleep(5 * 60)


def create_excel():
    wb = Workbook()  # create workbook
    ws = wb.active  # select worksheet
    ws.title = "Temperature"  # set worksheet title
    # set first row (titles) to bold
    row = ws.row_dimensions[1]
    row.font = Font(bold = True)
    # fill in titles
    ws.cell(row = 1, column = 1, value = "Day")
    ws.cell(row = 1, column = 2, value = "Hour")
    ws.cell(row = 1, column = 3, value = "Outside temperature (°C)")
    ws.cell(row = 1, column = 4, value = "Inside temperature (°C)")
    ws.cell(row = 1, column = 5, value = "Inside humidity (%)")

    # set column widths
    ws.column_dimensions['A'].width = 12  # Day
    ws.column_dimensions['B'].width = 10  # Hour
    ws.column_dimensions['C'].width = 10  # Outside temperature
    ws.column_dimensions['D'].width = 10  # Inside temperature
    ws.column_dimensions['E'].width = 10  # Inside humidity

    # save the file
    wb.save(__excel_file_name)

def write_to_excel(workbook, day, hour, outside_temp, inside_temp, inside_humid):
    sheet = workbook.active
    # get next unused row
    row = sheet.max_row + 1
    # write data to colums in row
    sheet.cell(row = row, column = 1, value = day)
    sheet.cell(row = row, column = 2, value = hour)
    sheet.cell(row = row, column = 3, value = outside_temp)
    sheet.cell(row = row, column = 4, value = inside_temp)
    sheet.cell(row = row, column = 5, value = inside_humid)
    # save the file
    workbook.save(__excel_file_name)

    print("Written to excel. Row {}".format(row))
    print("------------------------------------")


if __name__ == "__main__":
    try:
        # create an excel file if there is none
        if not os.path.isfile(__excel_file_name):
            create_excel()
            print("Excel log file created")
        else:
            print("Excel log file found")

        run()
    except KeyboardInterrupt:
        print("STOPPED")
